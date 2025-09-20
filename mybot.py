import requests
from bs4 import BeautifulSoup
from telegram import (
    Update,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    ReplyKeyboardRemove
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    CallbackQueryHandler,
)
import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import tempfile
import asyncio

BOT_TOKEN = "7757762485:AAHY5BrJ58YpdW50lAwRUsTwahtRDrd1RyA"
ADMIN_ID = 6550324099 

user_state = {}

STATS_FILE = "stats.json"
users_data = {}

# ====== Load stats ======
def load_stats():
    global users_data
    if not os.path.isfile(STATS_FILE):
        users_data = {}
        return
    try:
        with open(STATS_FILE, "r") as f:
            data = json.load(f)
            users_data = data if isinstance(data, dict) else {}
    except:
        users_data = {}

# ====== Save stats ======
def save_stats():
    try:
        with open(STATS_FILE, "w") as f:
            json.dump(users_data, f, indent=2)
    except:
        pass

# ====== Inline Keyboards ======
def get_main_inline_keyboard():
    keyboard = [
        [
            InlineKeyboardButton("🆓 Free Search", callback_data="free"),
            InlineKeyboardButton("💎 Premium Search", callback_data="premium"),
        ]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_free_inline_keyboard():
    keyboard = [
        [
            InlineKeyboardButton("📱 Search by Number", callback_data="search_number"),
            InlineKeyboardButton("🆔 Search by CNIC", callback_data="search_cnic"),
        ],
        [InlineKeyboardButton("⬅ Back", callback_data="back_main")]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_premium_inline_keyboard():
    keyboard = [
        [
            InlineKeyboardButton("🌳 Voter Tree", callback_data="premium_votertree"),
            InlineKeyboardButton("☎ PTCL Detail", callback_data="premium_ptcl"),
        ],
        [
            InlineKeyboardButton("📱 Number Ownership", callback_data="premium_number"),
            InlineKeyboardButton("🚗 Vehicle Detail", callback_data="premium_vehicle"),
        ],
        [
            InlineKeyboardButton("🆔 CNIC Detail", callback_data="premium_cnic"),
        ],
        [InlineKeyboardButton("⬅ Back", callback_data="back_main")]
    ]
    return InlineKeyboardMarkup(keyboard)

# ====== Fetch Voter Tree (dbfather.42web.io) ======
async def get_voter_tree(cnic: str) -> str:
    """
    Fetches HTML from dbfather index and parses the results table to return a formatted string.
    Runs blocking requests in a thread to avoid blocking the event loop.
    """
    try:
        def fetch_html():
            url = "https://dbfather.42web.io/index.php"
            payload = {"cnicInput": cnic}
            resp = requests.post(url, data=payload, timeout=30)
            return resp.text

        html = await asyncio.to_thread(fetch_html)
        soup = BeautifulSoup(html, "html.parser")

        table_div = soup.find("div", id="resultsTable")
        if not table_div:
            return "⚠️ No data found."

        table = table_div.find("table")
        if not table:
            return "⚠️ No result table found."

        tbody = table.find("tbody")
        if not tbody:
            return "⚠️ No family members found."

        rows = tbody.find_all("tr")
        if not rows:
            return "⚠️ No family members found."

        result_text = "👨‍👩‍👧‍👦 Family Tree Results:\n\n"
        for row in rows:
            cols = [td.get_text(strip=True) for td in row.find_all("td")]
            # expects: S.No, Name, CNIC, Age, Relation
            if len(cols) >= 5:
                result_text += (
                    f"👤 Name: {cols[1]}\n"
                    f"🆔 CNIC: {cols[2]}\n"
                    f"🎂 Age: {cols[3]}\n"
                    f"👥 Relation: {cols[4]}\n\n"
                )
        return result_text.strip()

    except Exception as e:
        return f"⚠️ Could not fetch data: {e}"

# ====== /start ======
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_id = str(user.id)
    username = user.username or user.first_name or "Unknown"

    if user_id not in users_data:
        users_data[user_id] = {"username": username, "search_count": 0, "searches": []}
        save_stats()
    else:
        users_data[user_id]["username"] = username
        save_stats()

    user_state.pop(update.effective_chat.id, None)
    await update.message.reply_text(
        "👋 Welcome! Please choose an option:",
        reply_markup=get_main_inline_keyboard()
    )

# ====== Callback Handler ======
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat_id

    # Free vs Premium
    if query.data == "free":
        await query.message.reply_text("🆓 Free Search Options:", reply_markup=get_free_inline_keyboard())
    elif query.data == "premium":
        await query.message.reply_text("💎 Premium Search Options:", reply_markup=get_premium_inline_keyboard())
    elif query.data == "back_main":
        await query.message.reply_text("⬅ Back to Main Menu:", reply_markup=get_main_inline_keyboard())

    # Free Search options
    elif query.data == "search_number":
        user_state[chat_id] = ("free", "number")
        await query.message.reply_text("📱 Please enter the mobile number (with 92, e.g., 923001234567):", reply_markup=ReplyKeyboardRemove())
    elif query.data == "search_cnic":
        user_state[chat_id] = ("free", "cnic")
        await query.message.reply_text("🆔 Please enter the CNIC number (13 digits):", reply_markup=ReplyKeyboardRemove())

    # Premium Search options
    elif query.data == "premium_votertree":
        user_state[chat_id] = ("premium", "votertree")
        await query.message.reply_text("🆔 Please enter the CNIC for Voter Tree:", reply_markup=ReplyKeyboardRemove())
    elif query.data.startswith("premium_"):
        await query.message.reply_text("❌ This premium feature is locked. Only Voter Tree is enabled.")

# ====== Handle searches ======
async def menu_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user = update.effective_user
    user_id = str(user.id)
    username = user.username or user.first_name or "Unknown"
    text = update.message.text.strip()

    if user_id not in users_data:
        users_data[user_id] = {"username": username, "search_count": 0, "searches": []}
    else:
        users_data[user_id]["username"] = username
    save_stats()

    if chat_id not in user_state:
        await update.message.reply_text("⚠ Please type /start and select an option.", reply_markup=get_main_inline_keyboard())
        return

    mode, search_type = user_state[chat_id]

    # ===== Free Search =====
    if mode == "free":
        if search_type == "number":
            # require country code 92 + 10 digits => length 12
            if not text.isdigit() or not text.startswith("92") or len(text) != 12:
                await update.message.reply_text("❌ Invalid number. Enter in format 92XXXXXXXXXX (e.g., 923001234567).")
                return
        elif search_type == "cnic":
            if not text.isdigit() or len(text) != 13:
                await update.message.reply_text("❌ Invalid CNIC. Enter exactly 13 digits.")
                return

        users_data[user_id]["search_count"] += 1
        users_data[user_id]["searches"].append({"type": search_type, "query": text})
        save_stats()

        await update.message.reply_text("🔍 Searching (Free)... Please wait.")

        # 🔄 Changed: Fetch from pakistandatabase.com
        try:
            def fetch_free():
                url = "https://pakistandatabase.com/index.php"
                payload = {"search_query": text}
                resp = requests.post(url, data=payload, timeout=30)
                return resp.text

            html = await asyncio.to_thread(fetch_free)
            soup = BeautifulSoup(html, "html.parser")

            table = soup.find("table", class_="api-response")
            if not table:
                await update.message.reply_text("⚠ No result found.")
                await send_developer_info(update)
                return

            tbody = table.find("tbody")
            if not tbody:
                await update.message.reply_text("⚠ No result found.")
                await send_developer_info(update)
                return

            rows = tbody.find_all("tr")
            result_text = ""
            for row in rows:
                cols = [col.get_text(strip=True) for col in row.find_all("td")]
                if len(cols) >= 4:
                    result_text += (
                        f"📱 Mobile: {cols[0]}\n"
                        f"👤 Name: {cols[1]}\n"
                        f"🆔 CNIC: {cols[2]}\n"
                        f"🏠 Address: {cols[3]}\n\n"
                    )

            await update.message.reply_text(result_text.strip() or "⚠ No data found.")
            await send_developer_info(update)

        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")
            await send_developer_info(update)

    # ===== Premium Search =====
    elif mode == "premium" and search_type == "votertree":
        if not text.isdigit() or len(text) != 13:
            await update.message.reply_text("❌ Invalid CNIC. Enter exactly 13 digits.")
            return

        users_data[user_id]["search_count"] += 1
        users_data[user_id]["searches"].append({"type": "votertree", "query": text})
        save_stats()

        await update.message.reply_text("🔍 Fetching Voter Tree... Please wait.")

        result = await get_voter_tree(text)

        # result is a formatted string already
        try:
            # try to send as markdown-safe text (we used plain chars so ok)
            await update.message.reply_text(result[:4000])
        except Exception:
            # fallback raw
            await update.message.reply_text(result[:4000])

        await send_developer_info(update)

# ===== Developer info ======
async def send_developer_info(update: Update):
    developer_msg = "🤖 Bot developed by Muazam Ali\n📞 WhatsApp: "
    await update.message.reply_text(developer_msg)
    await update.message.reply_text("Choose search type:", reply_markup=get_main_inline_keyboard())
    user_state.pop(update.effective_chat.id, None)

# ====== Stats Command ======
async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("⛔ Unauthorized.")
            return
        if not users_data:
            await update.message.reply_text("No users yet.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Stats"

        center = Alignment(horizontal="center", vertical="center")
        bold_font_white = Font(bold=True, color="FFFFFF")
        bold_font_black = Font(bold=True, color="000000")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        purple_fill = PatternFill(start_color="A47DB9", end_color="A47DB9", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        orange_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
        cyan_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

        row_num = 1

        for uid, data in users_data.items():
            username = data.get("username", "Unknown")
            search_count = data.get("search_count", 0)
            searches = data.get("searches", [])

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws.cell(row=row_num, column=1, value="User Name").fill = blue_fill
            ws.cell(row=row_num, column=1).alignment = center
            ws.cell(row=row_num, column=1).font = bold_font_white
            ws.cell(row=row_num, column=1).border = thin_border

            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=3, value="User Id").fill = purple_fill
            ws.cell(row=row_num, column=3).alignment = center
            ws.cell(row=row_num, column=3).font = bold_font_white
            ws.cell(row=row_num, column=3).border = thin_border

            row_num += 1

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws.cell(row=row_num, column=1, value=username).border = thin_border
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=3, value=uid).border = thin_border

            row_num += 1

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws.cell(row=row_num, column=1, value="Total Searches").fill = yellow_fill
            ws.cell(row=row_num, column=1).alignment = center
            ws.cell(row=row_num, column=1).font = bold_font_black
            ws.cell(row=row_num, column=1).border = thin_border

            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=3, value=search_count).fill = yellow_fill
            ws.cell(row=row_num, column=3).alignment = center
            ws.cell(row=row_num, column=3).font = bold_font_black
            ws.cell(row=row_num, column=3).border = thin_border

            row_num += 1

            ws.cell(row=row_num, column=1, value="SR").fill = green_fill
            ws.cell(row=row_num, column=2, value="Search Type").fill = orange_fill
            ws.cell(row=row_num, column=3, value="Search Query").fill = cyan_fill
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)

            for col in range(1, 5):
                ws.cell(row=row_num, column=col).alignment = center
                ws.cell(row=row_num, column=col).font = bold_font_white
                ws.cell(row=row_num, column=col).border = thin_border

            row_num += 1

            for idx, s in enumerate(searches, start=1):
                ws.cell(row=row_num, column=1, value=idx).fill = green_fill
                ws.cell(row=row_num, column=1).alignment = center

                ws.cell(row=row_num, column=2, value=s["type"]).fill = orange_fill
                ws.cell(row=row_num, column=2).alignment = center

                ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
                ws.cell(row=row_num, column=3, value=s["query"]).fill = cyan_fill
                ws.cell(row=row_num, column=3).alignment = center

                for col in range(1, 5):
                    ws.cell(row=row_num, column=col).border = thin_border

                row_num += 1

            row_num += 2

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)

        with open(tmp_path, "rb") as file:
            await context.bot.send_document(chat_id=ADMIN_ID, document=file, filename="user_stats.xlsx")

        os.remove(tmp_path)

    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")

# ====== Main ======
if __name__ == "__main__":
    load_stats()
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, menu_choice))
    print("🤖 Bot is running...")
    app.run_polling()
